"""Two-sheet .xlsx report meant to be shared with non-technical colleagues:
a "Results" sheet with the full per-question data (friendly columns like the
question/answers/judge verdict first, raw metrics like token counts after -
skip those if they don't mean anything to you), and a "Summary" sheet with
the same narrative text the CLI prints, so the whole story fits in one file.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

RESULTS_HEADERS = [
    "Question",
    "Full-Context Answer",
    "RAG Answer",
    "Judge Preferred",
    "Judge Reasoning",
    "Full-Context Latency (s)",
    "RAG Latency (s)",
    "Speed Ratio (FC / RAG)",
    "Judge: FC Accuracy",
    "Judge: FC Completeness",
    "Judge: FC Clarity",
    "Judge: RAG Accuracy",
    "Judge: RAG Completeness",
    "Judge: RAG Clarity",
    "FC Prompt Tokens",
    "FC Output Tokens",
    "FC Tokens/s",
    "RAG Prompt Tokens",
    "RAG Output Tokens",
    "RAG Tokens/s",
    "RAG Retrieved Pages",
]

# Column widths, aligned to RESULTS_HEADERS above (friendly columns wide, metrics narrow).
RESULTS_COLUMN_WIDTHS = [40, 60, 60, 16, 50, 12, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 16]

PREFERRED_LABELS = {"full_context": "Full-Context", "rag": "RAG", "tie": "Tie"}


def save_xlsx_report(path: str, fc_results, rag_results, judge_results, summary_lines: list[str]) -> None:
    wb = Workbook()

    results_sheet = wb.active
    results_sheet.title = "Results"
    _write_results_sheet(results_sheet, fc_results, rag_results, judge_results)

    summary_sheet = wb.create_sheet("Summary")
    _write_summary_sheet(summary_sheet, summary_lines)

    wb.save(path)


def _write_results_sheet(sheet, fc_results, rag_results, judge_results) -> None:
    judge_by_question = {jr.question: jr for jr in judge_results}

    sheet.append(RESULTS_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for fc, rag in zip(fc_results, rag_results):
        jr = judge_by_question.get(fc.question)
        rag_total_latency = rag.result.total_duration_s + rag.retrieval_s
        speed_ratio = fc.result.total_duration_s / rag_total_latency if rag_total_latency > 0 else None

        sheet.append([
            fc.question,
            fc.answer,
            rag.answer,
            PREFERRED_LABELS.get(jr.preferred, jr.preferred) if jr else "N/A (judge failed)",
            jr.reasoning if jr else None,
            round(fc.result.total_duration_s, 2),
            round(rag_total_latency, 2),
            round(speed_ratio, 2) if speed_ratio is not None else None,
            jr.full_context_scores["accuracy"] if jr else None,
            jr.full_context_scores["completeness"] if jr else None,
            jr.full_context_scores["clarity"] if jr else None,
            jr.rag_scores["accuracy"] if jr else None,
            jr.rag_scores["completeness"] if jr else None,
            jr.rag_scores["clarity"] if jr else None,
            fc.result.prompt_tokens,
            fc.result.output_tokens,
            round(fc.result.tokens_per_s, 1),
            rag.result.prompt_tokens,
            rag.result.output_tokens,
            round(rag.result.tokens_per_s, 1),
            ", ".join(str(p) for p in rag.retrieved_pages),
        ])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, width in enumerate(RESULTS_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    sheet.freeze_panes = "A2"


def _write_summary_sheet(sheet, summary_lines: list[str]) -> None:
    for line in summary_lines:
        sheet.append([line])
    sheet.column_dimensions["A"].width = 100
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
