from openpyxl import load_workbook

from ctxvsrag.approaches.full_context import ApiResult
from ctxvsrag.approaches.rag_answer import RagResult
from ctxvsrag.backends.base import ChatResult
from ctxvsrag.judge import JudgeResult
from ctxvsrag.report import RESULTS_HEADERS, save_xlsx_report


def _chat_result(**overrides) -> ChatResult:
    fields = dict(text="answer", prompt_tokens=100, output_tokens=20, total_duration_s=1.0)
    fields.update(overrides)
    return ChatResult(**fields)


def _fc_result(question: str, answer: str = "FC answer") -> ApiResult:
    return ApiResult(approach="full_context", question=question, answer=answer, result=_chat_result(text=answer))


def _rag_result(question: str, answer: str = "RAG answer") -> RagResult:
    return RagResult(
        approach="rag", question=question, answer=answer,
        retrieval_s=0.05, retrieved_pages=[1, 2], result=_chat_result(text=answer),
    )


def _judge_result(question: str, preferred: str = "full_context") -> JudgeResult:
    return JudgeResult(
        question=question,
        full_context_scores={"accuracy": 9, "completeness": 8, "clarity": 9},
        rag_scores={"accuracy": 7, "completeness": 6, "clarity": 7},
        preferred=preferred,
        reasoning="Full-context covered more detail.",
    )


def test_creates_two_sheets_named_results_and_summary(tmp_path):
    path = tmp_path / "report.xlsx"
    save_xlsx_report(str(path), [_fc_result("Q1?")], [_rag_result("Q1?")], [_judge_result("Q1?")], ["line one"])

    wb = load_workbook(path)
    assert wb.sheetnames == ["Results", "Summary"]


def test_results_sheet_has_expected_headers(tmp_path):
    path = tmp_path / "report.xlsx"
    save_xlsx_report(str(path), [_fc_result("Q1?")], [_rag_result("Q1?")], [_judge_result("Q1?")], ["line one"])

    wb = load_workbook(path)
    header_row = [cell.value for cell in wb["Results"][1]]
    assert header_row == RESULTS_HEADERS


def test_results_sheet_has_one_row_per_question(tmp_path):
    path = tmp_path / "report.xlsx"
    fc = [_fc_result("Q1?"), _fc_result("Q2?")]
    rag = [_rag_result("Q1?"), _rag_result("Q2?")]
    jr = [_judge_result("Q1?"), _judge_result("Q2?", preferred="rag")]
    save_xlsx_report(str(path), fc, rag, jr, ["line one"])

    wb = load_workbook(path)
    sheet = wb["Results"]
    assert sheet.max_row == 3  # header + 2 questions

    row1 = [cell.value for cell in sheet[2]]
    assert row1[0] == "Q1?"
    assert row1[1] == "FC answer"
    assert row1[2] == "RAG answer"
    assert row1[3] == "Full-Context"

    row2 = [cell.value for cell in sheet[3]]
    assert row2[3] == "RAG"


def test_missing_judge_result_shows_na_and_blank_scores(tmp_path):
    path = tmp_path / "report.xlsx"
    # No JudgeResult for this question - simulates a JudgeParseError that got skipped.
    save_xlsx_report(str(path), [_fc_result("Q1?")], [_rag_result("Q1?")], [], ["line one"])

    wb = load_workbook(path)
    row = [cell.value for cell in wb["Results"][2]]
    assert row[3] == "N/A (judge failed)"
    assert row[4] is None  # reasoning
    assert row[8] is None  # judge FC accuracy


def test_summary_sheet_contains_the_narrative_lines(tmp_path):
    path = tmp_path / "report.xlsx"
    lines = ["=" * 10, "SUMMARY", "Full-context preferred: 1"]
    save_xlsx_report(str(path), [_fc_result("Q1?")], [_rag_result("Q1?")], [_judge_result("Q1?")], lines)

    wb = load_workbook(path)
    summary_values = [row[0].value for row in wb["Summary"].iter_rows()]
    assert summary_values == lines


def test_speed_ratio_column_is_computed(tmp_path):
    path = tmp_path / "report.xlsx"
    fc = [_fc_result("Q1?")]
    fc[0].result.total_duration_s = 10.0
    rag = [_rag_result("Q1?")]
    rag[0].result.total_duration_s = 4.0
    rag[0].retrieval_s = 1.0  # total RAG latency = 5.0 -> ratio 10/5 = 2.0
    save_xlsx_report(str(path), fc, rag, [_judge_result("Q1?")], ["line one"])

    wb = load_workbook(path)
    row = [cell.value for cell in wb["Results"][2]]
    assert row[7] == 2.0  # "Speed Ratio (FC / RAG)" column
